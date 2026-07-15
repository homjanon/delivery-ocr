# -*- coding: utf-8 -*-
"""
送货单 → Excel 本地识别工具（纯本地运行，零网页、零 CORS）
====================================================================
用法:
  python ocr_delivery.py "图片文件夹/"      # 识别整个文件夹里的图片（批量）
  python ocr_delivery.py a.jpg b.jpg        # 指定若干张图片
  python ocr_delivery.py                     # 不带参数 → 弹文件选择框点选

依赖: requests openpyxl python-dotenv pillow(可选，用于压缩图片省 token)
配置: 同目录下放 .env 文件（参考 .env.example），填入 API Key 与 PROVIDER

说明:
  - 复用网页版 js/prompt.js 的 SYSTEM_PROMPT 与 js/excel.js 的表头/合计逻辑。
  - 模型调用改由本地 Python 发起，无浏览器 CORS 限制，因此商汤也可接（需按官方文档核对模型名）。
  - Win7 需使用 Python 3.9.x（3.10+ 已不支持 Win7）。详见 打包说明.md。
"""
import os
import sys
import re
import json
import time
import base64
import glob
from pathlib import Path

try:
    from dotenv import load_dotenv
except ImportError:
    def load_dotenv(*a, **k):
        return None

# ----- 定位程序基目录（脚本模式与 PyInstaller 打包后都正确）-----
BASE = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path(__file__).parent
load_dotenv(BASE / ".env")


# ============================ 复用网页版提示词（js/prompt.js 原样）============================
SYSTEM_PROMPT = """你是一名专业的送货单（出货单/发货单）识别专家。用户会提供一张或多张送货单图片。
请逐张、逐行识别每一项商品，输出一个 JSON 数组，数组中的每个对象代表一行商品，必须包含且仅包含以下字段（键名严格一致，禁止增减或改名）：
- "单据编号"：单号/编号（若无则空字符串）
- "日期"：开票日期（YYYY-MM-DD，若无则空字符串）
- "台头公司名称"：送货单台头（送货方）公司名称，位于图片正上方。同张单的多行此项相同。
- "名称及规格"：商品名称及规格描述
- "规格型号/颜色"：规格型号或颜色（若无则空字符串）
- "数量"：数值（数字，不含单位）
- "单位"：计量单位（如 台/个/箱/米/千克）
- "单价"：单价数值（数字，去除货币符号与千分位逗号）
- "金额"：金额数值（数字，去除货币符号与千分位逗号）
- "备注"：备注信息（若无则空字符串）
- "客户/收货单位"：客户或收货单位名称
要求：
1. 只输出 JSON 数组本身，不要任何解释、不要 Markdown 代码块标记、不要外围文字。
2. 所有金额、单价、数量必须是纯数字（如 1234.5），禁止出现 "¥"、"元"、逗号、空格等。
3. 若某字段缺失：文本字段填空字符串 ""，数量/单价/金额填 0。
4. 多张图片的商品合并到同一个数组中，按图片顺序、单据内行顺序排列。
5. 严格忠实于图片原文，不要推断或编造数据。"""


# ============================ 复用网页版表头（js/excel.js 原样）============================
EXCEL_HEADERS = [
    "单据编号", "日期", "台头公司名称", "名称及规格", "规格型号/颜色",
    "数量", "单位", "单价", "金额", "备注", "客户/收货单位",
]


# ============================ 模型预设（后端调用，无 CORS 限制）============================
# 智谱 / 硅基流动 已在网页版验证可用，参数一致。
# 商汤因无浏览器 CORS 限制，本地可接；模型名与返回格式请以官方文档为准（可用 SENSENOVA_MODEL 覆盖）。
MODEL_PRESETS = {
    "zhipu": {
        "name": "智谱 GLM-4.6V-Flash",
        "base_url": "https://open.bigmodel.cn/api/paas/v4/chat/completions",
        "model": "glm-4.6v-flash",
        "key_env": "ZHIPU_API_KEY",
    },
    "qwen35b": {
        "name": "硅基流动 Qwen3.5-35B-A3B",
        "base_url": "https://api.siliconflow.cn/v1/chat/completions",
        "model": "Qwen/Qwen3.5-35B-A3B",
        "key_env": "SILICONFLOW_API_KEY",
    },
    "sensenova": {
        "name": "商汤 SenseNova（需按官方文档核对模型名）",
        "base_url": "https://token.sensenova.cn/v1/chat/completions",
        "model": os.getenv("SENSENOVA_MODEL", "SenseNova-V6.7-Flash-Lite"),
        "key_env": "SENSENOVA_API_KEY",
    },
}


IMG_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp"}
MIME_MAP = {
    ".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
    ".bmp": "image/bmp", ".gif": "image/gif", ".webp": "image/webp",
}


def log(msg):
    print(msg, flush=True)


def collect_image_paths(args):
    """从命令行参数收集图片路径（支持文件夹）。"""
    paths = []
    for a in args:
        if os.path.isdir(a):
            for p in sorted(glob.glob(os.path.join(a, "*"))):
                if Path(p).suffix.lower() in IMG_EXTS:
                    paths.append(p)
        else:
            paths.append(a)
    return [p for p in paths if Path(p).suffix.lower() in IMG_EXTS]


def pick_files():
    """无参数时弹出文件选择框。"""
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError:
        log("未安装 tkinter，无法弹窗选择。请直接传入图片路径，例如：")
        log('  python ocr_delivery.py "图片文件夹/"')
        sys.exit(1)
    root = tk.Tk()
    root.withdraw()
    files = filedialog.askopenfilenames(
        title="选择送货单图片",
        filetypes=[("图片", "*.jpg *.jpeg *.png *.bmp *.gif *.webp")],
    )
    root.destroy()
    return list(files)


def compress_and_encode(path, max_dim=2000, quality=85):
    """读图→(可选压缩)→base64 data URL。无 Pillow 时退回原图字节。"""
    ext = Path(path).suffix.lower()
    mime = MIME_MAP.get(ext, "image/jpeg")
    try:
        from PIL import Image
        import io
        img = Image.open(path)
        w, h = img.size
        scale = min(1.0, max_dim / max(w, h))
        if scale < 1.0:
            img = img.resize((int(w * scale), int(h * scale)))
        buf = io.BytesIO()
        img.convert("RGB").save(buf, format="JPEG", quality=quality)
        data = buf.getvalue()
    except Exception:
        with open(path, "rb") as f:
            data = f.read()
    return "data:" + mime + ";base64," + base64.b64encode(data).decode("ascii")


def call_model(cfg, api_key, images, model_override=None):
    """调用视觉大模型，返回模型原始文本。含 429/502/503/504 重试。"""
    import requests

    model = model_override or cfg["model"]
    img_parts = [{"type": "image_url", "image_url": {"url": b}} for b in images]
    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": [
            {"type": "text", "text": "请识别以下送货单图片，严格按系统提示词输出 JSON 数组。"},
            *img_parts,
        ]},
    ]
    body = {"model": model, "messages": messages, "temperature": 0}
    body["max_tokens"] = 8192
    body["chat_template_kwargs"] = {"enable_thinking": False}

    headers = {"Content-Type": "application/json", "Authorization": "Bearer " + api_key}
    last_err = None
    for i in range(3):
        try:
            resp = requests.post(cfg["base_url"], headers=headers, json=body, timeout=(15, 240))
        except Exception as e:
            last_err = e
            if i < 2:
                time.sleep((i + 1) * 2)
                continue
            raise
        if resp.ok:
            data = resp.json()
            return data["choices"][0]["message"].get("content") \
                or data["choices"][0]["message"].get("reasoning") or ""
        if i < 2 and resp.status_code in (429, 502, 503, 504):
            wait = (i + 1) * 2
            log(f"服务繁忙（{resp.status_code}），{wait}秒后重试…")
            time.sleep(wait)
            last_err = resp
            continue
        raise RuntimeError(f"HTTP {resp.status_code} {resp.text[:300]}")
    raise RuntimeError(f"模型调用失败：{last_err}")


def extract_json(text):
    """从模型回复中提取 JSON 数组（复用网页版鲁棒逻辑）。"""
    t = (text or "").strip()
    t = re.sub(r"<think>[\s\S]*?<\/think>", "", t, flags=re.I)
    t = re.sub(r"^```(?:json)?\s*", "", t, flags=re.I)
    t = re.sub(r"```\s*$", "", t, flags=re.I)
    m = re.search(r"```(?:json)?\s*([\s\S]*?)```", t, flags=re.I)
    if m:
        t = m.group(1).strip()
    s = t.find("["); e = t.rfind("]")
    if s != -1 and e != -1 and e > s:
        t = t[s:e + 1]
    if not t:
        raise ValueError("模型未输出任何内容")
    if t.startswith("[") and not t.endswith("]"):
        t = re.sub(r",\s*$", "", t).rstrip() + "]"
    try:
        return json.loads(t)
    except Exception as err:
        raise ValueError("JSON 不完整/被截断（可能触及 max_tokens）：" + str(err))


def _num(v):
    try:
        return float(str(v).replace(",", ""))
    except Exception:
        return 0.0


def build_workbook(rows):
    """生成 xlsx：表头 + 数据 + 合计行（金额求和）+ 表头加粗。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "送货单"
    ws.append(EXCEL_HEADERS[:])

    for r in rows:
        ws.append([_norm(r.get(h), h) for h in EXCEL_HEADERS])

    amt_idx = EXCEL_HEADERS.index("金额")
    total = sum(_num(r.get("金额")) for r in rows)
    total_row = []
    for i, h in enumerate(EXCEL_HEADERS):
        if i == 0:
            total_row.append("合计")
        elif i == amt_idx:
            total_row.append(round(total * 100) / 100)
        else:
            total_row.append("")
    ws.append(total_row)

    fill = PatternFill("solid", fgColor="EAF1F5")
    for c in range(1, len(EXCEL_HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = 16
    return wb


def _norm(v, h):
    if h in ("数量", "单价", "金额"):
        return _num(v)
    return "" if v is None else v


def main():
    args = sys.argv[1:]
    if args:
        paths = collect_image_paths(args)
    else:
        paths = pick_files()

    if not paths:
        log("未选择任何图片，已退出。")
        sys.exit(0)

    provider = (os.getenv("PROVIDER") or "zhipu").strip().lower()
    if provider not in MODEL_PRESETS:
        log(f"PROVIDER 无效：{provider}（可选：{', '.join(MODEL_PRESETS)}）")
        sys.exit(1)
    cfg = MODEL_PRESETS[provider]
    api_key = os.getenv(cfg["key_env"])
    if not api_key:
        log(f"未找到 {cfg['key_env']}，请在 .env 中配置（参考 .env.example）。")
        sys.exit(1)

    # 多图时降低分辨率/质量，避免大图耗 token 或触发限流
    n = len(paths)
    max_dim = 1200 if n > 2 else 2000
    quality = 70 if n > 2 else 85
    log(f"压缩 {n} 张图片（max_dim={max_dim}, quality={quality}）…")
    images = [compress_and_encode(p, max_dim, quality) for p in paths]

    log(f"调用模型：{cfg['name']} …")
    t0 = time.time()
    raw = call_model(cfg, api_key, images)
    log(f"模型返回（前 200 字）：\n{raw[:200]}")
    rows = extract_json(raw)
    if not isinstance(rows, list):
        raise ValueError("返回不是数组")

    wb = build_workbook(rows)
    out = BASE / ("送货单_" + time.strftime("%Y-%m-%d") + ".xlsx")
    wb.save(out)
    log(f"完成：识别 {len(rows)} 行，耗时 {time.time() - t0:.1f}s")
    log(f"已保存：{out}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log("出错：" + str(e))
        sys.exit(1)
